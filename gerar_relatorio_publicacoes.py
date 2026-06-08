"""
gerar_relatorio_publicacoes.py — LegalOne · Publicações Diárias
Detecção de CNJ por nome (normalização robusta) e por conteúdo (fallback).
"""

import io, re, unicodedata
from datetime import date, datetime, timezone, timedelta
try:
    from zoneinfo import ZoneInfo as _ZoneInfo
    _BRASILIA = _ZoneInfo('America/Sao_Paulo')
except ImportError:
    try:
        from pytz import timezone as _tz
        _BRASILIA = _tz('America/Sao_Paulo')
    except ImportError:
        _BRASILIA = timezone(timedelta(hours=-3))  # UTC-3 fixo como último recurso
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


def normalizar(s):
    """Remove acentos e normaliza todos os tipos de espaço Unicode."""
    if not s or isinstance(s, float):
        return ""
    s = str(s).strip()
    for ch in '\xa0\u200b\u200c\u200d\u202f\u205f\u3000\ufeff':
        s = s.replace(ch, ' ')
    for code in range(0x2000, 0x200B):
        s = s.replace(chr(code), ' ')
    s = s.upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    while '  ' in s:
        s = s.replace('  ', ' ')
    return s.strip()


COORDENADORES_MAPEADOS = {
    # ── CAMILLA GOES BARBOSA ──────────────────────────────────────────────────
    normalizar("CAMILLA GOES BARBOSA"):                           "CAMILLA GOES BARBOSA",

    # ── GABRIEL GIORGIO CICCHELERO ───────────────────────────────────────────
    normalizar("GABRIEL GIORGIO CICCHELERO"):                     "GABRIEL GIORGIO CICCHELERO",
    normalizar("ALYSSON NARBAL DE OLIVEIRA SOMBRA"):              "GABRIEL GIORGIO CICCHELERO",
    normalizar("JAMILE BARRETO"):                                 "GABRIEL GIORGIO CICCHELERO",
    normalizar("JULIANA DE OLIVEIRA ROCHA"):                      "GABRIEL GIORGIO CICCHELERO",
    normalizar("RODRIGO RIBEIRO ANTUNES QUARIGUASI"):             "GABRIEL GIORGIO CICCHELERO",
    normalizar("DALILA DRISANA GOMES GONCALVES"):                 "GABRIEL GIORGIO CICCHELERO",
    normalizar("ANA VITORIA SALES DE OLIVEIRA FALCAO"):           "GABRIEL GIORGIO CICCHELERO",
    normalizar("RAFAEL CAVALCANTE BARBOSA"):                      "GABRIEL GIORGIO CICCHELERO",

    # ── HELANZIA DE ARAUJO XAVIER WICHAMNN ───────────────────────────────────
    normalizar("HELANZIA DE ARAUJO XAVIER WICHAMNN"):             "HELANZIA DE ARAUJO XAVIER WICHAMNN",

    # ── JENIFFER ROSA BARBOSA DE SALES ───────────────────────────────────────
    normalizar("JENIFFER ROSA BARBOSA DE SALES"):                 "JENIFFER ROSA BARBOSA DE SALES",
    normalizar("PAULO MARCIO SOARES DE CARVALHO FILHO"):          "JENIFFER ROSA BARBOSA DE SALES",

    # ── JULIANA MIRELLA ALVES RODRIGUES ──────────────────────────────────────
    normalizar("JULIANA MIRELLA ALVES RODRIGUES"):                "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("ARTHUR MASSARI"):                                 "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("DANIEL BARROS DE OLIVEIRA"):                      "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("GUSTAVO LOPES ALENCAR FILHO"):                    "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("KELIANE DE OLIVEIRA"):                            "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("MONIQUE DE KAROLIN SILVA DA COSTA"):              "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("NATALIA PAIVA DE PAULA"):                         "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("ROBERTA RAYANNE VASCONCELOS BOTO"):               "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("THALLYS ANDERSON FERREIRA DE LIMA"):              "JULIANA MIRELLA ALVES RODRIGUES",
    normalizar("VICTOR EMANOEL FRADIQUE ACCIOLY FONTENELE"):      "JULIANA MIRELLA ALVES RODRIGUES",

    # ── LUCIANE MODERNEL MENDES ───────────────────────────────────────────────
    normalizar("LUCIANE MODERNEL MENDES"):                        "LUCIANE MODERNEL MENDES",
    normalizar("ANTONIO EDUARDO GOES AGUIAR FILHO"):              "LUCIANE MODERNEL MENDES",
    normalizar("ERIKA PAULA SANTOS LIMA"):                        "LUCIANE MODERNEL MENDES",
    normalizar("SANE BORGES BORGOMONI"):                          "LUCIANE MODERNEL MENDES",

    # ── MARCELLE LEITE RENTROIA ───────────────────────────────────────────────
    normalizar("MARCELLE LEITE RENTROIA"):                        "MARCELLE LEITE RENTROIA",
    normalizar("MARIANA MOTA FROTA"):                             "MARCELLE LEITE RENTROIA",
    normalizar("YASMIM GORDIANO BARBOSA"):                        "MARCELLE LEITE RENTROIA",

    # ── NAYANDERSON LUAN MELLO PINHEIRO ──────────────────────────────────────
    normalizar("NAYANDERSON LUAN MELLO PINHEIRO"):                "NAYANDERSON LUAN MELLO PINHEIRO",
    normalizar("ANDRE VIANA GARRIDO"):                            "NAYANDERSON LUAN MELLO PINHEIRO",
    normalizar("EMERSON DE ALMEIDA MELO JUNIOR"):                 "NAYANDERSON LUAN MELLO PINHEIRO",
    normalizar("EMERSON TRAVASSOS TORQUATO"):                     "NAYANDERSON LUAN MELLO PINHEIRO",
    normalizar("JEAN VICTOR NUNES SARAIVA"):                      "NAYANDERSON LUAN MELLO PINHEIRO",

    # ── RONALD FEITOSA AGUIAR FILHO ───────────────────────────────────────────
    normalizar("RONALD FEITOSA AGUIAR FILHO"):                    "RONALD FEITOSA AGUIAR FILHO",
    normalizar("ALEXIA ALENCAR CAPIBARIBE"):                      "RONALD FEITOSA AGUIAR FILHO",

    # ── SUZANA MARIA CAMPOS MARANHAO DE LIMA ─────────────────────────────────
    normalizar("SUZANA MARIA CAMPOS MARANHAO DE LIMA"):           "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    normalizar("EVILANY GABRIELA BRAGA PONTES"):                  "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    normalizar("FRANCOISE CATHERINE SOUZA ALVES"):                "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    normalizar("GIOVANNA CAMPOS PEREIRA"):                        "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    normalizar("MATHEUS CAVALCANTI DE ARAUJO"):                   "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    normalizar("TATIANE CARMO SANTA ROSA"):                       "SUZANA MARIA CAMPOS MARANHAO DE LIMA",

    # ── TICIANNA PIRES DE SOUZA ───────────────────────────────────────────────
    normalizar("TICIANNA PIRES DE SOUZA"):                        "TICIANNA PIRES DE SOUZA",

    # ── YURI ALVES BARROS DOS SANTOS ─────────────────────────────────────────
    normalizar("YURI ALVES BARROS DOS SANTOS"):                   "YURI ALVES BARROS DOS SANTOS",
    normalizar("JULIA MENEZES MORGADO"):                    "YURI ALVES BARROS DOS SANTOS",
    normalizar("LUIZ GUILHERME GONCALVES GIRAO"):                 "YURI ALVES BARROS DOS SANTOS",
}
COORDENADORES_CONHECIDOS = [
    "CAMILLA GOES BARBOSA",
    "GABRIEL GIORGIO CICCHELERO",
    "HELANZIA DE ARAUJO XAVIER WICHAMNN",
    "JENIFFER ROSA BARBOSA DE SALES",
    "JULIANA MIRELLA ALVES RODRIGUES",
    "LUCIANE MODERNEL MENDES",
    "MARCELLE LEITE RENTROIA",
    "NAYANDERSON LUAN MELLO PINHEIRO",
    "RONALD FEITOSA AGUIAR FILHO",
    "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "TICIANNA PIRES DE SOUZA",
    "YURI ALVES BARROS DOS SANTOS",
]

INATIVOS = {normalizar(n) for n in [
    "ALEXANDRE SOUZA DA SILVA FILHO","ALICE RIBEIRO MELO CRISOSTOMO",
    "ALLY CHARLYS MUSSE MENDES FILHO","AMANDA BEZERRA DE MENEZES NETO",
    "AMANDA MOURA DOS SANTOS BRAGA","ANA CAMILA CIFONI DE VASCONCELOS BARROSO",
    "ANA CAROLINA PAES GALVAO DE MELO SANTOS","ANA KATHARINE VASCONCELOS DE SOUSA",
    "ANA LIA TERCEIRO ALMEIDA","ANA LUIZA ROCHA PICANCO","ANA PAULA MILEO",
    "ANA VITORIA SALES DE OLIVEIRA FALCAO","ANDRE LUIS SILVA DE SOUZA",
    "ANDRESSA HELENA ANTUNES DE OLIVEIRA","ANNA KAROLINE CARDOSO AQUINO",
    "ANNE AGUIAR BARBOSA","APARECIDO JOSE SILVA PLATERO","ARTHUR LEMOS DE AGUIAR",
    "ARTUR SARAIVA DE ANDRADE","BRENO LIMA PITTA PINHEIRO","BRUNA DOS SANTOS CASTRO",
    "BRUNO MACHADO FORTES","CAIO CESAR PINHEIRO GUERREIRO",
    "CARMEN GEORGIA REBOUCAS DE OLIVEIRA JORGE VIEIRA",
    "CICERO WELLINGTON BATISTA DO NASCIMENTO","CLAYTON SANTANA LUZ",
    "CRISTIANE SANTOS DOS REIS","CRISTIANO HOLANDA DA CUNHA",
    "DALILA CARLOS DE CASTRO","DALILA DRISANA GOMES GONCALVES",
    "DALVA REGINA MARTINS ARGENTAO","DANIEL ROCHA FERREIRA EUGENIO",
    "DANIEL SARAIVA PINHEIRO","DANIELA MONDINO CANTORI","DANIELLE DE QUEIROZ ALVES",
    "DAVID BARROSO PEREIRA","DEBORA MARIA TEIXEIRA AUGUSTO LIMA",
    "DENISE ARAUJO SILVA DOS SANTOS","DIEGO LIMA HOLANDA DOS SANTOS",
    "DIOGO DE SOUZA ROSA","EDUARDO HENRIQUE AGUIAR","ELANE GERMANO NUNES ALVES",
    "EMANUELLY ARAUJO VIEIRA","EMERSON DE ALMEIDA MELO JUNIOR",
    "EMILLY TEIXEIRA DE SOUSA","ERICA VALESCA DE OLIVEIRA FRAGA",
    "FABIO AUGUSTO SILVERIO SILVA","FABIOLA FARIAS IBIAPINA",
    "FELIPE AGUIAR DE NEGREIROS ANDRADE","FELIPE BEZERRA RIBEIRO",
    "FELIPE CARVALHO CARNEIRO","FELIPE DE ALBUQUERQUE BEZERRA",
    "FERNANDA LIMA ALVES","FERNANDA SOARES RODRIGUES","FERNANDO GARCIA",
    "FERRUCIO ALISON ALCANTARA AMORIM","FLAVIA PESSOA MONTEIRO",
    "FRAN COSTA DE CASTRO","FRANCISCA EDILANDIA FREITAS ARAUJO",
    "FRANCISCA THAYSSE LIMA COSTA","FRANCISCO ROBERTO DE MATOS",
    "GABRIEL CORREA FURTADO","GISELE PEREIRA FONTELES",
    "HELIDA ZEDNIK RODRIGUES LIMA","HELOISA GOMES REBOUCAS","IAGO AMARAL REIS",
    "IGOR RABELO MAGALHAES","IPRAZOS","IRANEIVA ROCHA QUIRINO BARROS",
    "IURY ALVES VIEIRA DE SOUSA","JAMILE DE GOIS RODRIGUES AMORIM",
    "JANAINA ELIAS CHIARADIA","JANE DIANE DE RAMOS NUNES GONCALVES",
    "JEAN VICTOR NUNES SARAIVA","JOAO CARNEIRO MELLO MOREIRA",
    "JOAO MARCOS DE ABREU TEIXEIRA","JOHANN DANIEL DE OLIVEIRA INOCENCIO",
    "JOSE DE ARIMATEIA GORDIANO OLIVEIRA BARBOSA","JOSE ZITO RABELO NETO",
    "JOSEPH MICAIAS OLIVEIRA DE CASTRO","JULHIERME ALEX ZANAQUI",
    "JULIANA MARA VASCONCELOS ANDRADE","JULIANA OSELAME MACEDO ZANAQUI",
    "JULIANA VECCHI DA SILVA","KAMILA BARBOSA OLIVEIRA","KASSIA FAUSTINO COELHO",
    "LAURO LINHARES LEITE","LAYLA EVELYN NASCIMENTO PINHEIRO",
    "LEANDRO SARUBBI DE CARVALHO ROCHA","LETICIA DE ALCANTARA MENDES",
    "LETICIA LEMOS BARRETO","LEVY MOTA DE OLIVEIRA","LUAN PEDRO CIANFARANI",
    "LUANA PRESTES DE SOUSA MELO","LUCAS ANDRADE NOBREGA","LUCIANO PEROBA FILHO",
    "LUIZ GUILHERME GONCALVES GIRAO","MACSIMUS WALESKO DE CASTRO DUARTE",
    "MANOEL VICTOR BACALHAU","MARCELO LIMA ARRAIS","MARIA EDUARDA SANTOS ARAUJO",
    "MARIA LAURA MELO ALMEIDA","MARIA THAIS RODRIGUES DA COSTA",
    "MARIANA ALMEIDA DE SOUZA","MARIANA FASANARO DE CARVALHO",
    "MARIANA SILVA DE MORAIS","MARILIA BARROSO WALRAVEN CUNHA",
    "MARINA NOBRE SIMAO","MARIO FONSECA GOMES FERREIRA",
    "MATHEUS ARRUDA ALBUQUERQUE","MATHEUS HAYASAKI",
    "MONICA MARIA QUEIROZ DA SILVA","MONICA SANTOS MARTINS",
    "OYSTR - ROBO","PEDRO ELIAS STELMACHUK COSTA",
    "PEDRO HENRIQUE ROOSEVELT GOES BARBOSA","PHAMELLA SALES DE SOUZA COSTA",
    "RAFAEL PEREIRA DE SOUZA","RAFAELA JOZIA HOLANDA",
    "RAUL MATIAS DA SILVA PADRAO","RAYANNE LARI SOUSA ALMEIDA",
    "REBECA RODRIGUES DO NASCIMENTO","REBECCA ARAUJO ROSA",
    "REGINALDO DE BRITO OLIVEIRA JUNIOR","RENAN PEREIRA DA SILVA",
    "RENAN SINHORINI FUSATO","ROBERTA FURTADO DE ARRAES ALENCAR E CASTRO",
    "RODRIGO RIBEIRO ANTUNES QUARIGUASI","ROGERIO SCARABEL BARBOSA",
    "RUAN PEREIRA DO NASCIMENTO","RUBIANA APARECIDA BARBIERI",
    "SAMARA DE MOURA FERREIRA","SAMYA MONTEIRO PAMPLONA DE OLIVEIRA",
    "SANDRA MARA DO NASCIMENTO","SARAH CHRISTINE ROCHA LOBAO",
    "SCHEILA DE PAULA CORDEIRO","SUZANA MARIA LIMA BARROSO FELIX",
    "TAMIRIS CAMELO MELO LINO","TATIANA HAUBERT","THALYTA MARIA TORQUATO VITOR",
    "THIAGO GURGEL FREIRE LEITE","VICTOR MARTINS BARBOSA",
    "VICTORIA FERREIRA ALONSO","VITORIA CAVALCANTE DOS SANTOS",
    "VITORIA DA SILVA FREITAS","WALESSA DIOGENES PEIXOTO DE ALENCAR",
    "WANDERLUCY CORREIA DE ALMEIDA","WELLINGTON PEREIRA DA ROCHA FILHO",
    "WELLINGTON RUBENS","WESLLEY MACEDO DE OLIVEIRA","WVENDEL SENA OLIVEIRA",
    "YASMIM GORDIANO BARBOSA",
]}

OUTPUT_COLS = [
    ('Data cadastro', 20, ['Data cadastro','Data Cadastro','Data']),
    ('Pasta',         16, ['Pasta','PASTA']),
    ('Natureza',      14, ['Natureza','NATUREZA','natureza']),
    ('Responsavel',   32, ['Responsavel','Responsável','RESPONSAVEL']),
    ('Status',        12, ['Status','STATUS','status']),
    ('Cliente',       30, ['Cliente','CLIENTE','cliente']),
    ('Numero de CNJ', 28, ['Numero de CNJ','Número de CNJ','CNJ']),
]


def _encontrar_coluna(df, candidatos):
    mapa = {normalizar(c): c for c in df.columns}
    for c in candidatos:
        k = normalizar(c)
        if k in mapa:
            return mapa[k]
    for c in candidatos:
        k = normalizar(c)
        for col_norm, col_orig in mapa.items():
            if k in col_norm or col_norm in k:
                return col_orig
    return None


def _encontrar_cnj(df):
    """Por nome (normalizado) ou por padrão de conteúdo NNNNNNN-NN.NNNN.N.NN.NNNN."""
    col = _encontrar_coluna(df, ['Numero de CNJ','Número de CNJ','CNJ','Nº de CNJ'])
    if col:
        return col
    cnj_pat = re.compile(r'^\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}$')
    for col in df.columns:
        amostra = df[col].dropna().head(10).astype(str).str.strip()
        if amostra.apply(lambda v: bool(cnj_pat.match(v))).sum() >= 3:
            return col
    return None


def _encontrar_resp(df):
    """Detecta coluna Responsável por nome ou por conteúdo (nomes conhecidos)."""
    col = _encontrar_coluna(df, ['Responsavel', 'Responsável', 'RESPONSAVEL',
                                  'Responsavel pelo processo', 'Nome responsavel'])
    if col:
        return col
    # Fallback: procura coluna cujos valores batem com nomes conhecidos
    conhecidos = set(list(INATIVOS)[:30])  # amostra dos nomes conhecidos
    for col in df.columns:
        try:
            amostra = df[col].dropna().head(20).astype(str)
            hits = amostra.apply(lambda v: normalizar(v) in conhecidos).sum()
            if hits >= 2:
                return col
        except Exception:
            pass
    # Último recurso: coluna de string com valores longos e multi-palavras (nomes)
    melhor_col, melhor_score = None, 0
    for col in df.columns:
        try:
            amostra = df[col].dropna().head(20).astype(str)
            score = amostra.apply(
                lambda v: len(v) > 10 and ' ' in v.strip() and v.strip().replace(' ','').isalpha()
            ).sum()
            if score > melhor_score:
                melhor_score, melhor_col = score, col
        except Exception:
            pass
    return melhor_col if melhor_score >= 3 else None


def _carregar_df(data):
    # Guardar bytes brutos para poder reler múltiplas vezes
    raw = data if isinstance(data, bytes) else data.read()

    KW = {'NATUREZA','RESPONSAVEL','CLIENTE','PASTA','STATUS','CADASTRO','CNJ','NUMERO'}

    def _header_ok(df):
        return any(any(kw in normalizar(str(c)) for kw in KW) for c in df.columns)

    def _primeira_linha_e_cabecalho(df):
        if df.empty: return False
        return any(any(kw in normalizar(str(v)) for kw in KW) for v in df.iloc[0])

    for header_row in [0, 1, 2]:
        try:
            df = pd.read_excel(io.BytesIO(raw), header=header_row)
            if _header_ok(df) and not _primeira_linha_e_cabecalho(df):
                return df
        except Exception:
            pass

    # Fallback
    try:
        return pd.read_excel(io.BytesIO(raw))
    except Exception:
        return pd.read_excel(io.BytesIO(raw), engine='openpyxl')


def _extrair_data(df, filename=""):
    m = re.search(r'(\d{2})[_.\-/](\d{2})[_.\-/](\d{2,4})', filename)
    if m:
        try:
            y = int(m.group(3))
            if y < 100:
                y += 2000
            d = date(y, int(m.group(2)), int(m.group(1)))
            return d, d.strftime('%d/%m/%Y')
        except Exception:
            pass
    # Fallback: usa a data atual no fuso de Brasília (o Streamlit Cloud roda em UTC).
    # A coluna "Data cadastro" NÃO é usada pois contém data de cadastro, não de emissão.
    today = datetime.now(_BRASILIA).date()
    return today, today.strftime('%d/%m/%Y')


def _dia_norm(d):
    return ['segunda','terca','quarta','quinta','sexta','sabado','domingo'][d.weekday()]


def _dia_pt(d):
    return ['Segunda-feira','Terça-feira','Quarta-feira','Quinta-feira',
            'Sexta-feira','Sábado','Domingo'][d.weekday()]


def classificar_tribunal(cnj):
    if not cnj or pd.isna(cnj):
        return 'OUTROS'
    p = str(cnj).strip().split('.')
    if len(p) >= 3:
        j = p[2].strip()
        if j == '5': return 'TRT'
        if j == '4': return 'TRF'
        if j == '8': return 'TJ'
    return 'OUTROS'


def _fill(h):  return PatternFill(fill_type='solid', start_color=h, end_color=h)
def _f(bold=False, size=9, color='FF000000', italic=False):
    return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)
def _b():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def _a(h='left', v='center'):
    return Alignment(horizontal=h, vertical=v)
def _c(ws, row, col, val, font=None, fill=None, align=None, border=True):
    cell = ws.cell(row=row, column=col)
    cell.value = val
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = _b()
    return cell


FT = _fill('FF1F3864')  # título
FH = _fill('FF2E4D8A')  # header coluna
FM = _fill('FFFF8C00')  # tribunal minoritário
FI = _fill('FFF4A460')  # inativo
FD = _fill('FFF0F4FF')  # dados fundo claro
FW = _fill('FFFFFFFF')  # linha alternada
FG = _fill('FFEEEEEE')  # ausente
FR = _fill('FF8B0000')  # divisão especial

AF = {
    'VANESSA':     _fill('FFB3D9FF'),
    'PALOMA':      _fill('FFC8F7C5'),
    'BARBARA':     _fill('FFFFE4B5'),
    'ANNA JULIA':  _fill('FFFFD6D6'),
    'ANA CECILIA': _fill('FFFFE4F4'),
    'TATIANA':     _fill('FFFFF0C0'),
}

GC = ['FF1F3864','FF2E4D8A','FF3A5FA8','FF4A72C4','FF5A85D8','FF6B98EC','FF7BABFF','FF8CBEFD','FF9DCFFE']


def _grad(n, s='FF1F3864', e='FFA2D4FF'):
    def h2(h): h=h[2:]; return int(h[:2],16),int(h[2:4],16),int(h[4:],16)
    def r2(r,g,b): return 'FF%02X%02X%02X'%(r,g,b)
    if n<=1: return [s]
    sr,sg,sb=h2(s); er,eg,eb=h2(e)
    return [r2(int(sr+(er-sr)*i/(n-1)),int(sg+(eg-sg)*i/(n-1)),int(sb+(eb-sb)*i/(n-1))) for i in range(n)]


def _calcular_cotas(n, dia, analistas_excluidos=None):
    if analistas_excluidos is None: analistas_excluidos = []
    aj     = dia != 'quinta' and 'ANNA JULIA' not in analistas_excluidos
    full   = [a for a in ['VANESSA','PALOMA','BARBARA'] if a not in analistas_excluidos]
    half   = [a for a in ['ANA CECILIA'] if a not in analistas_excluidos]
    if aj:     half.append('ANNA JULIA')
    div_st = len(full) + len(half)*0.5
    q_st   = n/div_st if div_st else 0
    tat    = q_st > 40
    full   = full + (['TATIANA'] if tat else [])
    div    = len(full) + len(half)*0.5
    q      = int(n/div) if div else 0
    qh     = q//2
    cotas  = {a: q for a in full}
    for a in half: cotas[a] = qh
    left   = n - (len(full)*q + len(half)*qh)
    for a in full:
        if left <= 0: break
        cotas[a] += 1; left -= 1
    return cotas


def _distribuir(df, cotas, col_cnj, col_nat, col_cli):
    alloc = {a: [] for a in cotas}
    pool  = list(df.index)
    if col_cli and 'BARBARA' in cotas:
        gpm  = [i for i in pool if normalizar(str(df.at[i,col_cli])).startswith('GPM')]
        take = gpm[:cotas['BARBARA']]
        alloc['BARBARA'].extend(take)
        pool = [i for i in pool if i not in set(take)]
    if col_nat:
        nlt = {i for i in pool
               if normalizar(str(df.at[i,col_nat])) not in {'TRABALHISTA'}
               and str(df.at[i,col_nat]).strip()}
        nl  = [i for i in pool if i in nlt]
        for a in ['ANA CECILIA','ANNA JULIA','PALOMA']:
            if a not in cotas: continue
            need = cotas[a] - len(alloc[a])
            take = nl[:need]
            alloc[a].extend(take)
            nl   = nl[need:]
            pool = [i for i in pool if i not in set(take)]
    for a in ['VANESSA','PALOMA','BARBARA','TATIANA','ANNA JULIA','ANA CECILIA']:
        if a not in cotas or not pool: continue
        need = cotas[a] - len(alloc[a])
        if need > 0:
            alloc[a].extend(pool[:need])
            pool = pool[need:]
    if pool and 'VANESSA' in alloc:
        alloc['VANESSA'].extend(pool)
    return alloc


NOMES = {'VANESSA':'VANESSA','PALOMA':'PALOMA','BARBARA':'BÁRBARA',
         'ANNA JULIA':'ANNA JÚLIA','ANA CECILIA':'ANA CECÍLIA','TATIANA':'TATIANA'}
ORDEM = ['VANESSA','PALOMA','BARBARA','ANNA JULIA','ANA CECILIA','TATIANA']


def _build_resumo(ws, d_str, dia_nome, total_bruto, total_unico, cotas, alloc,
                  dia, df, col_nat, col_cli, especial=False):
    for col,w in [(1,20),(2,14),(3,12),(4,38),(5,16)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.merge_cells('A1:E1')
    _c(ws,1,1,('⚠️ DIVISÃO ESPECIAL — ' if especial else '')+f'RELATÓRIO DE PUBLICAÇÕES — {d_str} ({dia_nome})',
       _f(True,14,'FFFFFFFF'), FR if especial else FT, _a('center'))
    ws.row_dimensions[1].height = 32

    n_gpm   = sum(1 for i in df.index if col_cli and normalizar(str(df.at[i,col_cli])).startswith('GPM')) if col_cli else 0
    n_ntrab = sum(1 for i in df.index if col_nat and normalizar(str(df.at[i,col_nat])) not in {'TRABALHISTA'} and str(df.at[i,col_nat]).strip()) if col_nat else 0

    for i,(lbl,val) in enumerate([('Total bruto',total_bruto),('Publicações GPM',n_gpm),
                                   ('Duplicatas removidas',total_bruto-total_unico),('Não-trabalhistas',n_ntrab),
                                   ('Total único',total_unico)]):
        r,c = 2+i//2, 1 if i%2==0 else 3
        _c(ws,r,c,lbl,_f(True,9),FD,_a()); _c(ws,r,c+1,val,_f(size=9),align=_a('center'))

    r_nat = 5
    if col_nat:
        nc = df[col_nat].value_counts()
        for ci,lbl in enumerate(['DISTRIBUIÇÃO POR NATUREZA','TOTAL','%'],1):
            _c(ws,r_nat,ci,lbl,_f(True,9,'FFFFFFFF'),FH,_a('center'))
        for j,(nat,cnt) in enumerate(nc.items(),1):
            rr = r_nat+j; f = FD if j%2==1 else FW
            _c(ws,rr,1,str(nat),_f(size=9),f,_a())
            _c(ws,rr,2,int(cnt),_f(size=9),f,_a('center'))
            _c(ws,rr,3,f'{cnt/total_unico*100:.1f}%' if total_unico else '0%',_f(size=9),f,_a('center'))
        r_tab = r_nat+len(nc)+2
    else:
        r_tab = 7

    for ci,lbl in enumerate(['ANALISTA','PUBLICAÇÕES','%','OBSERVAÇÃO','STATUS'],1):
        _c(ws,r_tab,ci,lbl,_f(True,10,'FFFFFFFF'),FH,_a('center'))
    ws.row_dimensions[r_tab].height = 22

    for i,a in enumerate(ORDEM):
        r    = r_tab+1+i
        fill = AF.get(a,FG)
        if a=='ANNA JULIA' and dia=='quinta':
            vals=['—','—','','⛔ Fora (quinta-feira)']; fill=FG
        elif a not in cotas:
                             vals=['-','-','','⚪ Fora (cota ≤ 40)' if a=='TATIANA' else '']; fill=FG
        else:
            qtd  = len(alloc.get(a,[]))
            pct  = f'{qtd/total_unico*100:.1f}%' if total_unico else '0%'
            obs  = 'Prioridade GPM' if a=='BARBARA' else \
                   'Prioridade não-trabalhista (50%)' if a in ('ANNA JULIA','ANA CECILIA') else \
                   'Prioridade não-trabalhista' if a=='PALOMA' else ''
            stat = '✅ Incluída' if a=='TATIANA' else '50% da cota' if a in ('ANNA JULIA','ANA CECILIA') else '🟢 Ativa'
            vals = [qtd,pct,obs,stat]
        row_vals = [NOMES.get(a,a)] + (vals if len(vals)==4 else [*vals,''])
        for ci,v in enumerate(row_vals,1):
            _c(ws,r,ci,v,_f(size=9),fill,_a('center' if ci>1 else 'left'))


def _add_pendencias_resumo(ws, unmapped, inativos, r_start):
    """Adiciona tabela de responsáveis sem coordenador no RESUMO."""
    if not unmapped:
        return r_start
    r = r_start + 1
    # Título da seção
    ws.merge_cells(f'A{r}:E{r}')
    _c(ws, r, 1,
       '⚠️ RESPONSÁVEIS SEM COORDENADOR — Informar ao gestor do sistema',
       _f(True, 10, 'FFFFFFFF'), _fill('FF8B0000'), _a('center'))
    ws.row_dimensions[r].height = 20
    r += 1
    # Cabeçalho
    for ci, lbl in enumerate(['RESPONSÁVEL', 'STATUS NO SISTEMA', 'AÇÃO NECESSÁRIA'], 1):
        _c(ws, r, ci, lbl, _f(True, 9, 'FFFFFFFF'), FH, _a('center'))
    ws.merge_cells(f'C{r}:E{r}')
    r += 1
    for i, resp in enumerate(sorted(unmapped)):
        inativo = normalizar(resp) in INATIVOS
        status  = '⚠️ Inativo' if inativo else 'Ativo'
        fill    = _fill('FFFCE4D6') if inativo else (FD if i%2==0 else FW)
        _c(ws, r, 1, resp,   _f(size=9), fill, _a())
        _c(ws, r, 2, status, _f(size=9), fill, _a('center'))
        ws.merge_cells(f'C{r}:E{r}')
        _c(ws, r, 3, 'Coordenador não mapeado — informar ao gestor do sistema',
           _f(italic=True, size=9, color='FF8B0000'), fill, _a())
        r += 1
    return r


def _build_analista(ws, analista, nome, rows_df, df_cols, d_str):
    if rows_df.empty: return
    col_map  = {normalizar(c): c for c in df_cols}

    def find_src(cands):
        for c in cands:
            k = normalizar(c)
            if k in col_map: return col_map[k]
        for c in cands:
            k = normalizar(c)
            for cn,co in col_map.items():
                if k in cn or cn in k: return co
        return None

    out = [(d, w, find_src(cands)) for d,w,cands in OUTPUT_COLS]
    # Fallback: se nenhuma coluna foi encontrada, usar colunas reais do df (exceto CJ)
    if all(s is None for _,_,s in out):
        excluir = {'CJ'}
        out = [(c, 18, c) for c in rows_df.columns if normalizar(c) not in excluir]
    n   = len(out)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    _c(ws,1,1,f'{nome} — {d_str}',_f(True,11,'FFFFFFFF'),FT,_a('center'))
    ws.row_dimensions[1].height = 24
    for ci,(d,w,_) in enumerate(out,1):
        _c(ws,2,ci,d,_f(True,10,'FFFFFFFF'),FH,_a('center'))
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A3'

    cnj_src = next((src for _,_,src in out if src and 'CNJ' in normalizar(src)), None)
    if cnj_src and cnj_src in rows_df.columns:
        tribs = rows_df[cnj_src].apply(classificar_tribunal)
    else:
        tribs = pd.Series(['OUTROS']*len(rows_df), index=rows_df.index)

    ct     = tribs.value_counts()
    grupos = [g for g in ct.index if g != 'OUTROS']
    minT   = ct.loc[grupos].idxmin() if len(grupos) > 1 else None
    fb     = AF.get(analista, FD)

    for ri,(idx,_) in enumerate(rows_df.iterrows(), start=3):
        t   = tribs[idx] if idx in tribs.index else 'OUTROS'
        f   = FM if t==minT else (FW if (ri-2)%2==0 else fb)
        fc  = 'FFFFFFFF' if t==minT else 'FF000000'
        for ci,(_,_,src) in enumerate(out,1):
            val = rows_df.at[idx,src] if src and src in rows_df.columns else ''
            _c(ws,ri,ci,val,_f(size=9,color=fc),f,_a())

    if minT:
        leg = len(rows_df)+6
        ws.merge_cells(start_row=leg,start_column=1,end_row=leg,end_column=n)
        _c(ws,leg,1,f'🟠 Laranja = tribunal minoritário — {minT} ({ct.get(minT,0)} publ.)',
           _f(italic=True,size=8,color='FF7B2D00'),None,_a(),border=False)


def _build_coord(ws, df, col_resp, mapeamento, d_str, total):
    for col,w in [(1,42),(2,16),(3,14)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.merge_cells('A1:C1')
    _c(ws,1,1,f'TOTAL POR COORDENADOR — {d_str}',_f(True,13,'FFFFFFFF'),FT,_a('center'))
    ws.row_dimensions[1].height = 30
    for ci,lbl in enumerate(['COORDENADOR','PUBLICAÇÕES','% DO TOTAL'],1):
        _c(ws,2,ci,lbl,_f(True,10,'FFFFFFFF'),FH,_a('center'))

    ct = {}
    if col_resp:
        for v in df[col_resp].dropna():
            coord_raw = mapeamento.get(normalizar(str(v).strip()))
            if coord_raw is None:
                continue  # responsável sem nenhum mapeamento — omitir
            ct[coord_raw] = ct.get(coord_raw, 0) + 1

    dados = sorted(ct.items(), key=lambda x: -x[1])
    for i,(c,q) in enumerate(dados,3):
        h = GC[min(i-3,len(GC)-1)]
        pct = f'{q/total*100:.1f}%' if total else '0%'
        for ci,val in enumerate([c,q,pct],1):
            _c(ws,i,ci,val,_f(size=9,color='FFFFFFFF'),_fill(h),_a('center' if ci>1 else 'left'))
    # Linha de TOTAL GERAL — usa o total do dia (parâmetro), não só os mapeados
    tot_row = len(dados) + 3
    _c(ws,tot_row,1,'TOTAL GERAL',_f(bold=True,size=9,color='FFFFFFFF'),FT,_a('left'))
    _c(ws,tot_row,2,total,        _f(bold=True,size=9,color='FFFFFFFF'),FT,_a('center'))
    _c(ws,tot_row,3,'100%' if total else '0%',_f(bold=True,size=9,color='FFFFFFFF'),FT,_a('center'))
    if dados:
        n     = len(dados)
        chart = BarChart(); chart.type='bar'; chart.style=10
        chart.title='Publicações por Coordenador'
        chart.width=22; chart.height=14
        chart.add_data(Reference(ws,min_col=2,min_row=2,max_row=2+n),titles_from_data=True)
        chart.set_categories(Reference(ws,min_col=1,min_row=3,max_row=2+n))
        chart.series[0].graphicalProperties.solidFill='2E4D8A'
        ws.add_chart(chart,'D3')

def _build_resp(ws, df, col_resp, mapeamento, d_str, total):
    for col,w in [(1,38),(2,36),(3,14),(4,12),(5,14)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.merge_cells('A1:E1')
    _c(ws,1,1,f'TOTAL POR RESPONSÁVEL — {d_str}',_f(True,13,'FFFFFFFF'),FT,_a('center'))
    ws.row_dimensions[1].height = 30
    for ci,lbl in enumerate(['RESPONSÁVEL','COORDENADOR','STATUS','TOTAL','%'],1):
        _c(ws,2,ci,lbl,_f(True,10,'FFFFFFFF'),FH,_a('center'))
    # Conta publicações por responsável, agrupando variações de grafia pelo nome normalizado
    ct_norm    = {}  # norm_key -> count
    ct_display = {}  # norm_key -> nome original para exibição
    if col_resp:
        for v in df[col_resp].dropna():
            raw = str(v).strip()
            nk  = normalizar(raw)
            if not nk: continue
            ct_norm[nk]    = ct_norm.get(nk, 0) + 1
            if nk not in ct_display:
                ct_display[nk] = raw  # guarda o primeiro nome visto para exibição
    ct    = {ct_display[k]: ct_norm[k] for k in ct_norm}
    dados = sorted(ct.items(), key=lambda x: -x[1])
    cores = _grad(max(len(dados),1))
    for i,(resp,q) in enumerate(dados,3):
        norm   = normalizar(resp)
        coord  = mapeamento.get(norm,'—')
        inativ = norm in INATIVOS
        pct    = f'{q/total*100:.1f}%' if total else '0%'
        ff     = FI if inativ else _fill(cores[min(i-3,len(cores)-1)])
        fc     = 'FF000000' if inativ else 'FFFFFFFF'
        for ci,val in enumerate([resp,coord,'⚠️ Inativo' if inativ else 'Ativo',q,pct],1):
            _c(ws,i,ci,val,_f(size=9,color=fc),ff,_a('center' if ci>2 else 'left'))
    # Linha de TOTAL GERAL
    tot_r = len(dados) + 3
    total_resp = sum(q for _,q in dados)
    pct_tot = '100%' if total else '0%'
    ws.merge_cells(start_row=tot_r,start_column=1,end_row=tot_r,end_column=3)
    _c(ws,tot_r,1,'TOTAL GERAL',_f(bold=True,size=9,color='FFFFFFFF'),FT,_a('left'))
    _c(ws,tot_r,4,total_resp,   _f(bold=True,size=9,color='FFFFFFFF'),FT,_a('center'))
    _c(ws,tot_r,5,pct_tot,      _f(bold=True,size=9,color='FFFFFFFF'),FT,_a('center'))
    leg = len(dados)+4
    ws.merge_cells(start_row=leg,start_column=1,end_row=leg,end_column=5)
    _c(ws,leg,1,'⚠️ Laranja claro = Responsável inativo no sistema',
       _f(italic=True,size=8,color='FF7B2D00'),None,_a(),border=False)
    nc = min(15,len(dados))
    if nc:
        ch = BarChart(); ch.type='bar'; ch.style=10; ch.title='Top 15 Responsáveis'
        ch.width=24; ch.height=16
        ch.add_data(Reference(ws,min_col=4,min_row=2,max_row=2+nc),titles_from_data=True)
        ch.set_categories(Reference(ws,min_col=1,min_row=3,max_row=2+nc))
        ch.series[0].graphicalProperties.solidFill='3A5FA8'
        ws.add_chart(ch,'G3')


def pre_check(input_data, filename="", extra_mappings=None):
    if extra_mappings is None: extra_mappings = {}
    mp  = {**COORDENADORES_MAPEADOS, **{normalizar(k):v for k,v in extra_mappings.items()}}
    df  = _carregar_df(input_data)
    cc  = _encontrar_cnj(df)
    cr  = _encontrar_resp(df)
    tot = len(df)
    if cc: df = df.drop_duplicates(subset=[cc])
    uniq = len(df)
    d, ds = _extrair_data(df, filename)
    SKIP = {normalizar(c) for c in [
        'Responsavel','Responsável','RESPONSAVEL','Data cadastro','Natureza',
        'Cliente','Status','Pasta','Numero de CNJ','Número de CNJ','CNJ','CJ','Nome',
    ]}
    un, in_ = [], []
    if cr:
        for v in df[cr].dropna().unique():
            n = normalizar(str(v))
            if n in SKIP or len(str(v).strip()) < 3: continue
            if n in INATIVOS: in_.append(str(v).strip())
            if n not in mp:   un.append(str(v).strip())
    return {"data":d,"data_str":ds,"dia_semana":_dia_norm(d),"dia_semana_nome":_dia_pt(d),
            "total_bruto":tot,"total_unico":uniq,"duplicatas":tot-uniq,
            "unmapped":sorted(set(un)),"inativos_encontrados":sorted(set(in_))}


def gerar_relatorio(input_data, filename="", extra_mappings=None, divisao_especial=False, analistas_excluidos=None):
    if extra_mappings is None: extra_mappings = {}
    mp  = {**COORDENADORES_MAPEADOS, **{normalizar(k):v for k,v in extra_mappings.items()}}
    raw = input_data if isinstance(input_data,bytes) else input_data.read()
    df  = _carregar_df(raw)
    cc  = _encontrar_cnj(df)
    cn  = _encontrar_coluna(df,['Natureza','NATUREZA','natureza'])
    cli = _encontrar_coluna(df,['Cliente','CLIENTE','cliente'])
    cr  = _encontrar_resp(df)
    tot = len(df)
    if cc: df = df.drop_duplicates(subset=[cc])
    df  = df.reset_index(drop=True)
    uniq = len(df)
    d, ds = _extrair_data(df, filename)
    dia   = _dia_norm(d)
    if analistas_excluidos is None: analistas_excluidos = []
    cotas = _calcular_cotas(uniq, dia, analistas_excluidos)
    alloc = _distribuir(df, cotas, cc, cn, cli)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet('RESUMO')
    _build_resumo(ws, ds, _dia_pt(d), tot, uniq, cotas, alloc, dia, df, cn, cli, divisao_especial)
    # Calcular não mapeados e adicionar ao RESUMO
    NOMES_COLUNA2 = {normalizar(c) for c in [
        'Responsável','Responsavel','RESPONSAVEL','Responsavel pelo processo',
        'Data cadastro','Data Cadastro','Natureza','Cliente','Status','Pasta',
        'Numero de CNJ','Número de CNJ','CNJ','CJ','Data','Nome',
    ]}
    _unmapped = []
    if cr:
        for val in df[cr].dropna().unique():
            norm2 = normalizar(str(val))
            if norm2 in NOMES_COLUNA2: continue
            if len(str(val).strip()) < 3: continue
            if norm2 not in mp:
                _unmapped.append(str(val).strip())
    _r_pendencias = ws.max_row + 2
    _add_pendencias_resumo(ws, sorted(set(_unmapped)), list(INATIVOS), _r_pendencias)
    for a in ORDEM:
        if a in alloc and alloc[a]:
            _build_analista(wb.create_sheet(NOMES[a]), a, NOMES[a], df.loc[alloc[a]], df.columns.tolist(), ds)
    _build_coord(wb.create_sheet('POR COORDENADOR'), df, cr, mp, ds, uniq)
    _build_resp(wb.create_sheet('POR RESPONSÁVEL'), df, cr, mp, ds, uniq)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue(), {
        "data_str":ds,"dia_semana_nome":_dia_pt(d),"total_bruto":tot,"total_unico":uniq,
        "duplicatas":tot-uniq,"cotas":cotas,"tatiana_incluida":'TATIANA' in cotas,
        "anna_julia_ativa":dia!='quinta',
        "alloc_counts":{a:len(v) for a,v in alloc.items()},
    }
    
